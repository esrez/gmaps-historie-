"""Kniha jízd – generování jízd z historie polohy, ruční úpravy, pravidla
kilometrů podle místa, roční tachometr a export ve formátu XLSX vhodném
pro import do programu SPZ (Milk Computers).

Pozn. k SPZ: program importuje knihu jízd z xlsx; vozidlo (SPZ) musí být
v programu SPZ založené, jinak import odmítne. Prázdný řidič se doplní
z karty vozidla.

Pojmy:
- private  = soukromá jízda firemním vozidlem (v knize zůstává, jen označená)
- excluded = jízda soukromým autem (v knize jízd ani v exportu se neobjeví)
"""
from __future__ import annotations

import json
import math
import os
import time
from contextlib import closing

from fastapi import APIRouter, HTTPException, Query
from pydantic import BaseModel

from . import db, places
from .common import fmt_dt, haversine_m, local_dt, ts_range, xlsx_response

router = APIRouter(prefix="/api/trips")

# typy aktivit považované za jízdu autem (normalizované)
CAR_TYPES = {"IN_PASSENGER_VEHICLE", "DRIVING", "MOTORCYCLING", "IN_VEHICLE"}

SEMANTIC_CZ = places.SEMANTIC_CZ   # sdílený překlad Home/Work → Domov/Práce


def _norm(s: str | None) -> str:
    return (s or "").strip().lower()


def _round_km(km: float, round_up: bool) -> float:
    return float(math.ceil(km - 1e-9)) if round_up else round(km, 1)


class PlaceNamer:
    """Pojmenuje souřadnici podle nejbližšího navštíveného místa z historie.

    Pojmenovaná místa se načtou jedním dotazem předem – generování stovek
    jízd pak nedělá dotaz na každou souřadnici zvlášť.
    """

    def __init__(self, conn, radius_m: float = 350):
        self.radius_m = radius_m
        self.cache: dict = {}
        self.custom = places.load_places(conn)   # vlastní názvy mají přednost
        self.places = conn.execute(
            "SELECT COALESCE(NULLIF(name,''), semantic, address) label, semantic, "
            "       AVG(lat) lat, AVG(lon) lon "
            "FROM visits WHERE name IS NOT NULL OR semantic IS NOT NULL "
            "   OR address IS NOT NULL "
            "GROUP BY label, semantic").fetchall()

    def name(self, lat, lon) -> str:
        if lat is None or lon is None:
            return ""
        key = (round(lat, 4), round(lon, 4))
        if key in self.cache:
            return self.cache[key]
        custom = places.custom_label(self.custom, lat, lon)
        if custom:
            self.cache[key] = custom
            return custom
        best, best_d = None, self.radius_m + 1
        for p in self.places:
            if abs(p["lat"] - lat) > 0.01 or abs(p["lon"] - lon) > 0.02:
                continue
            d = haversine_m(lat, lon, p["lat"], p["lon"])
            if d < best_d:
                label = SEMANTIC_CZ.get((p["semantic"] or "").upper()) \
                    if p["label"] == p["semantic"] else p["label"]
                if label:
                    best, best_d = label, d
        result = best or f"{lat:.4f}, {lon:.4f}"
        self.cache[key] = result
        return result


def _load_rules(conn) -> list[dict]:
    return [dict(r) for r in conn.execute(
        "SELECT id, origin, destination, km FROM km_rules ORDER BY destination, origin")]


def _rule_km(rules: list[dict], origin: str, destination: str) -> float | None:
    """Km podle pravidla: pár odkud+kam platí obousměrně, pravidlo bez
    „odkud" platí pro všechny jízdy na dané místo (tam i zpět)."""
    o, d = _norm(origin), _norm(destination)
    best = None
    for r in rules:
        ro, rd = _norm(r["origin"]), _norm(r["destination"])
        if ro:
            if (ro, rd) in ((o, d), (d, o)):
                return r["km"]          # přesný pár má přednost
        elif rd and rd in (o, d):
            best = r["km"]
    return best


# ------------------------------------------------------------ modely

class GenerateParams(BaseModel):
    from_ts: int | None = None
    to_ts: int | None = None
    workdays_only: bool = True
    hour_from: int = 6
    hour_to: int = 18
    min_km: float = 0.5
    round_up: bool = True
    purpose: str = "Služební jízda"
    driver: str = ""
    plate: str = ""


class TripIn(BaseModel):
    start_ts: int
    end_ts: int
    km: float = 0
    origin: str = ""
    destination: str = ""
    purpose: str = ""
    driver: str = ""
    plate: str = ""
    private: bool = False
    excluded: bool = False


class TripPatch(BaseModel):
    start_ts: int | None = None
    end_ts: int | None = None
    km: float | None = None
    origin: str | None = None
    destination: str | None = None
    purpose: str | None = None
    driver: str | None = None
    plate: str | None = None
    private: bool | None = None
    excluded: bool | None = None


class PropagateParams(BaseModel):
    trip_id: int
    km: float
    from_ts: int | None = None
    to_ts: int | None = None
    round_up: bool = True
    save_rule: bool = True


class RuleIn(BaseModel):
    origin: str = ""
    destination: str
    km: float


class OdometerIn(BaseModel):
    year: int
    km: float
    plate: str = ""


def _row(r) -> dict:
    d = dict(r)
    d["private"] = bool(d["private"])
    d["excluded"] = bool(d["excluded"])
    return d


def _plate_sql(plate: str | None) -> tuple[str, tuple]:
    """Volitelný filtr vozidla (prázdné = všechna)."""
    if plate:
        return " AND LOWER(TRIM(COALESCE(plate,''))) = ?", (plate.strip().lower(),)
    return "", ()


TRIP_COLS = ("id", "start_ts", "end_ts", "km", "origin", "destination",
             "purpose", "driver", "plate", "private", "activity_ts", "excluded")


UNDO_DEPTH = 10


def _save_undo(conn, op: str, rows=(), created=()):
    """Uloží podklad pro vrácení hromadné akce (drží se posledních 10 kroků)."""
    payload = json.dumps(
        {"rows": [dict(r) for r in rows], "created": [int(c) for c in created]},
        ensure_ascii=False)
    conn.execute("INSERT INTO undo_log(created, op, data) VALUES(?,?,?)",
                 (int(time.time()), op, payload))
    conn.execute(
        "DELETE FROM undo_log WHERE id NOT IN "
        "(SELECT id FROM undo_log ORDER BY id DESC LIMIT ?)", (UNDO_DEPTH,))


# --------------------------------------------------------------- jízdy

@router.get("")
def list_trips(from_ts: int | None = Query(None), to_ts: int | None = Query(None),
               plate: str | None = Query(None)):
    lo, hi = ts_range(from_ts, to_ts)
    psql, pargs = _plate_sql(plate)
    with closing(db.connect()) as conn:
        rows = conn.execute(
            f"SELECT * FROM trips WHERE start_ts BETWEEN ? AND ?{psql} ORDER BY start_ts",
            (lo, hi, *pargs)).fetchall()
    return {"trips": [_row(r) for r in rows],
            "total_km": round(sum(r["km"] for r in rows if not r["excluded"]), 1)}


@router.post("/generate")
def generate(p: GenerateParams):
    lo, hi = ts_range(p.from_ts, p.to_ts)
    created = 0
    with closing(db.connect()) as conn:
        acts = conn.execute(
            "SELECT start_ts, end_ts, distance_m, start_lat, start_lon, end_lat, end_lon "
            "FROM activities WHERE start_ts BETWEEN ? AND ? "
            f"AND REPLACE(UPPER(type), ' ', '_') IN ({','.join('?' * len(CAR_TYPES))}) "
            "ORDER BY start_ts",
            (lo, hi, *CAR_TYPES)).fetchall()
        namer = PlaceNamer(conn)
        rules = _load_rules(conn)
        # ochrana proti duplicitám: intervaly už existujících jízd + nově přijatých
        covered = [(r["start_ts"], r["end_ts"]) for r in conn.execute(
            "SELECT start_ts, end_ts FROM trips WHERE start_ts BETWEEN ? AND ?", (lo, hi))]
        created_ids: list[int] = []
        skipped_dup = 0
        for a in acts:
            km = (a["distance_m"] or 0) / 1000
            if km < p.min_km:
                continue
            local = local_dt(a["start_ts"])
            if p.workdays_only and local.weekday() >= 5:
                continue
            if not (p.hour_from <= local.hour < p.hour_to):
                continue
            dur = max(a["end_ts"] - a["start_ts"], 1)
            if any(min(e, a["end_ts"]) - max(s, a["start_ts"]) > 0.5 * dur
                   for s, e in covered):
                skipped_dup += 1   # stejná cesta už v knize je (např. z druhého exportu)
                continue
            origin = namer.name(a["start_lat"], a["start_lon"])
            destination = namer.name(a["end_lat"], a["end_lon"])
            rule_km = _rule_km(rules, origin, destination)
            km_final = _round_km(rule_km if rule_km is not None else km, p.round_up)
            cur = conn.execute(
                "INSERT OR IGNORE INTO trips(start_ts, end_ts, km, origin, destination,"
                " purpose, driver, plate, private, activity_ts)"
                " VALUES(?,?,?,?,?,?,?,?,0,?)",
                (a["start_ts"], a["end_ts"], km_final, origin, destination,
                 p.purpose, p.driver, p.plate, a["start_ts"]))
            if cur.rowcount:
                created += 1
                created_ids.append(cur.lastrowid)
                covered.append((a["start_ts"], a["end_ts"]))
        if created_ids:
            _save_undo(conn, "generate", created=created_ids)
        conn.commit()
    return {"created": created, "scanned": len(acts), "skipped_duplicates": skipped_dup}


@router.post("")
def create_trip(t: TripIn):
    with closing(db.connect()) as conn:
        cur = conn.execute(
            "INSERT INTO trips(start_ts, end_ts, km, origin, destination,"
            " purpose, driver, plate, private, excluded) VALUES(?,?,?,?,?,?,?,?,?,?)",
            (t.start_ts, t.end_ts, t.km, t.origin, t.destination,
             t.purpose, t.driver, t.plate, int(t.private), int(t.excluded)))
        conn.commit()
        row = conn.execute("SELECT * FROM trips WHERE id=?", (cur.lastrowid,)).fetchone()
    return _row(row)


@router.post("/propagate")
def propagate(p: PropagateParams):
    """Po zadání km jedné jízdě doplní stejné km všem jízdám na stejné trase
    (obousměrně) ve zvoleném období a volitelně uloží trvalé pravidlo."""
    lo, hi = ts_range(p.from_ts, p.to_ts)
    with closing(db.connect()) as conn:
        trip = conn.execute("SELECT * FROM trips WHERE id=?", (p.trip_id,)).fetchone()
        if trip is None:
            raise HTTPException(404, "Jízda nenalezena")
        o, d = _norm(trip["origin"]), _norm(trip["destination"])
        if not d and not o:
            raise HTTPException(400, "Jízda nemá vyplněné odkud/kam")
        km = _round_km(p.km, p.round_up)
        affected = conn.execute(
            "SELECT * FROM trips WHERE start_ts BETWEEN ? AND ? AND excluded=0 "
            "AND ((LOWER(TRIM(COALESCE(origin,'')))=? AND LOWER(TRIM(COALESCE(destination,'')))=?) "
            " OR  (LOWER(TRIM(COALESCE(origin,'')))=? AND LOWER(TRIM(COALESCE(destination,'')))=?))",
            (lo, hi, o, d, d, o)).fetchall()
        _save_undo(conn, "propagate", rows=affected)
        cur = conn.execute(
            "UPDATE trips SET km=? WHERE start_ts BETWEEN ? AND ? AND excluded=0 "
            "AND ((LOWER(TRIM(COALESCE(origin,'')))=? AND LOWER(TRIM(COALESCE(destination,'')))=?) "
            " OR  (LOWER(TRIM(COALESCE(origin,'')))=? AND LOWER(TRIM(COALESCE(destination,'')))=?))",
            (km, lo, hi, o, d, d, o))
        updated = cur.rowcount
        if p.save_rule and (o or d):
            conn.execute(
                "INSERT INTO km_rules(origin, destination, km) VALUES(?,?,?) "
                "ON CONFLICT(origin, destination) DO UPDATE SET km=excluded.km",
                (trip["origin"] or "", trip["destination"] or "", km))
        conn.commit()
        rows = conn.execute(
            "SELECT * FROM trips WHERE start_ts BETWEEN ? AND ? ORDER BY start_ts",
            (lo, hi)).fetchall()
    return {"updated": updated, "km": km,
            "trips": [_row(r) for r in rows],
            "total_km": round(sum(r["km"] for r in rows if not r["excluded"]), 1)}


@router.post("/apply_rules")
def apply_rules(from_ts: int | None = Query(None), to_ts: int | None = Query(None),
                round_up: bool = Query(True)):
    """Přepočítá km všech nevyřazených jízd v období podle uložených pravidel."""
    lo, hi = ts_range(from_ts, to_ts)
    updated = 0
    with closing(db.connect()) as conn:
        rules = _load_rules(conn)
        rows = conn.execute(
            "SELECT * FROM trips "
            "WHERE start_ts BETWEEN ? AND ? AND excluded=0", (lo, hi)).fetchall()
        _save_undo(conn, "apply_rules", rows=rows)
        for r in rows:
            rule_km = _rule_km(rules, r["origin"] or "", r["destination"] or "")
            if rule_km is None:
                continue
            km = _round_km(rule_km, round_up)
            if km != r["km"]:
                conn.execute("UPDATE trips SET km=? WHERE id=?", (km, r["id"]))
                updated += 1
        conn.commit()
    return {"updated": updated}


@router.patch("/{trip_id}")
def update_trip(trip_id: int, patch: TripPatch):
    fields = patch.model_dump(exclude_unset=True)
    if not fields:
        raise HTTPException(400, "Žádná pole ke změně")
    for flag in ("private", "excluded"):
        if flag in fields:
            fields[flag] = int(fields[flag])
    sets = ", ".join(f"{k}=?" for k in fields)
    with closing(db.connect()) as conn:
        cur = conn.execute(f"UPDATE trips SET {sets} WHERE id=?",
                           (*fields.values(), trip_id))
        conn.commit()
        if cur.rowcount == 0:
            raise HTTPException(404, "Jízda nenalezena")
        row = conn.execute("SELECT * FROM trips WHERE id=?", (trip_id,)).fetchone()
    return _row(row)


@router.delete("/{trip_id}")
def delete_trip(trip_id: int):
    with closing(db.connect()) as conn:
        cur = conn.execute("DELETE FROM trips WHERE id=?", (trip_id,))
        conn.commit()
    if cur.rowcount == 0:
        raise HTTPException(404, "Jízda nenalezena")
    return {"deleted": trip_id}


@router.delete("")
def delete_range(from_ts: int | None = Query(None), to_ts: int | None = Query(None),
                 plate: str | None = Query(None)):
    """Smaže všechny jízdy ve zvoleném období (např. před novým generováním)."""
    lo, hi = ts_range(from_ts, to_ts)
    psql, pargs = _plate_sql(plate)
    with closing(db.connect()) as conn:
        rows = conn.execute(
            f"SELECT * FROM trips WHERE start_ts BETWEEN ? AND ?{psql}",
            (lo, hi, *pargs)).fetchall()
        _save_undo(conn, "delete_range", rows=rows)
        cur = conn.execute(
            f"DELETE FROM trips WHERE start_ts BETWEEN ? AND ?{psql}", (lo, hi, *pargs))
        conn.commit()
    return {"deleted": cur.rowcount}


@router.get("/undo")
def undo_info():
    with closing(db.connect()) as conn:
        steps = conn.execute("SELECT COUNT(*) c FROM undo_log").fetchone()["c"]
        row = conn.execute("SELECT created, op, data FROM undo_log "
                           "ORDER BY id DESC LIMIT 1").fetchone()
    if row is None:
        return {"available": False, "steps": 0}
    data = json.loads(row["data"])
    return {"available": True, "steps": steps, "op": row["op"],
            "created": row["created"],
            "affected": len(data["rows"]) + len(data["created"])}


@router.post("/undo")
def undo_last():
    """Vrátí poslední hromadnou akci (generování, propagaci km, použití
    pravidel nebo smazání období)."""
    with closing(db.connect()) as conn:
        row = conn.execute("SELECT id, op, data FROM undo_log "
                           "ORDER BY id DESC LIMIT 1").fetchone()
        if row is None:
            raise HTTPException(404, "Není co vracet")
        data = json.loads(row["data"])
        for cid in data["created"]:
            conn.execute("DELETE FROM trips WHERE id=?", (cid,))
        for r in data["rows"]:
            conn.execute(
                f"INSERT OR REPLACE INTO trips({','.join(TRIP_COLS)}) "
                f"VALUES({','.join('?' * len(TRIP_COLS))})",
                tuple(r.get(c) for c in TRIP_COLS))
        conn.execute("DELETE FROM undo_log WHERE id=?", (row["id"],))
        conn.commit()
    return {"op": row["op"], "restored": len(data["rows"]),
            "removed": len(data["created"])}


@router.get("/missing_days")
def missing_days(from_ts: int | None = Query(None), to_ts: int | None = Query(None),
                 workdays_only: bool = Query(True), min_km: float = Query(0.5)):
    """Dny, kdy podle historie proběhla jízda autem, ale v knize jízd chybí.
    Slouží jako upozornění na nevykázané dny."""
    lo, hi = ts_range(from_ts, to_ts)
    with closing(db.connect()) as conn:
        acts = conn.execute(
            "SELECT a.start_ts, a.distance_m FROM activities a "
            "LEFT JOIN trips t ON t.activity_ts = a.start_ts "
            "WHERE a.start_ts BETWEEN ? AND ? AND t.id IS NULL "
            f"AND REPLACE(UPPER(a.type), ' ', '_') IN ({','.join('?' * len(CAR_TYPES))})",
            (lo, hi, *CAR_TYPES)).fetchall()
        trip_days = {r["d"] for r in conn.execute(
            "SELECT DISTINCT date(start_ts,'unixepoch','localtime') d FROM trips "
            "WHERE start_ts BETWEEN ? AND ?", (lo, hi))}
    days: dict[str, dict] = {}
    for a in acts:
        km = (a["distance_m"] or 0) / 1000
        if km < min_km:
            continue
        local = local_dt(a["start_ts"])
        if workdays_only and local.weekday() >= 5:
            continue
        key = local.date().isoformat()
        if key in trip_days:
            continue   # den už v knize nějaké jízdy má
        d = days.setdefault(key, {"date": key, "km": 0.0, "count": 0})
        d["km"] = round(d["km"] + km, 1)
        d["count"] += 1
    return {"days": sorted(days.values(), key=lambda x: x["date"])}


@router.get("/alerts")
def alerts(from_ts: int | None = Query(None), to_ts: int | None = Query(None)):
    """Upozornění pro knihu jízd: neúplné jízdy a překročený tachometr."""
    lo, hi = ts_range(from_ts, to_ts)
    with closing(db.connect()) as conn:
        incomplete = conn.execute(
            "SELECT COUNT(*) c FROM trips WHERE start_ts BETWEEN ? AND ? AND excluded=0 "
            "AND (km <= 0 OR COALESCE(TRIM(destination),'') = '')", (lo, hi)).fetchone()["c"]
        odo_over = []
        for r in conn.execute("SELECT year, km FROM odometer"):
            booked = conn.execute(
                "SELECT COALESCE(SUM(km),0) s FROM trips WHERE excluded=0 "
                "AND strftime('%Y', start_ts, 'unixepoch', 'localtime') = ?",
                (str(r["year"]),)).fetchone()["s"]
            if booked > r["km"]:
                odo_over.append({"year": r["year"], "odometer_km": r["km"],
                                 "booked_km": round(booked, 1)})
    return {"incomplete_trips": incomplete, "odometer_exceeded": odo_over}


# ------------------------------------------------------------- pravidla

@router.get("/rules")
def list_rules():
    with closing(db.connect()) as conn:
        return {"rules": _load_rules(conn)}


@router.post("/rules")
def upsert_rule(r: RuleIn):
    if not r.destination.strip() and not r.origin.strip():
        raise HTTPException(400, "Vyplňte alespoň cíl (kam)")
    with closing(db.connect()) as conn:
        conn.execute(
            "INSERT INTO km_rules(origin, destination, km) VALUES(?,?,?) "
            "ON CONFLICT(origin, destination) DO UPDATE SET km=excluded.km",
            (r.origin.strip(), r.destination.strip(), r.km))
        conn.commit()
        return {"rules": _load_rules(conn)}


@router.delete("/rules/{rule_id}")
def delete_rule(rule_id: int):
    with closing(db.connect()) as conn:
        cur = conn.execute("DELETE FROM km_rules WHERE id=?", (rule_id,))
        conn.commit()
    if cur.rowcount == 0:
        raise HTTPException(404, "Pravidlo nenalezeno")
    return {"deleted": rule_id}


# ------------------------------------------------------------ tachometr

@router.get("/odometer")
def get_odometer(year: int = Query(...), plate: str = Query("")):
    """Roční nájezd dle tachometru vs. kilometry vykázané v knize jízd.
    Tachometr je vedený zvlášť pro každé vozidlo (SPZ; prázdná = společný)."""
    plate = plate.strip()
    psql, pargs = _plate_sql(plate)
    with closing(db.connect()) as conn:
        row = conn.execute("SELECT km FROM odometer WHERE year=? AND plate=?",
                           (year, plate)).fetchone()
        booked = conn.execute(
            f"SELECT COALESCE(SUM(km),0) s FROM trips WHERE excluded=0 "
            f"AND strftime('%Y', start_ts, 'unixepoch', 'localtime') = ?{psql}",
            (str(year), *pargs)).fetchone()["s"]
    odo = row["km"] if row else None
    return {"year": year, "plate": plate, "odometer_km": odo,
            "booked_km": round(booked, 1),
            "remaining_km": round(odo - booked, 1) if odo is not None else None}


@router.put("/odometer")
def set_odometer(o: OdometerIn):
    with closing(db.connect()) as conn:
        conn.execute(
            "INSERT INTO odometer(year, plate, km) VALUES(?,?,?) "
            "ON CONFLICT(year, plate) DO UPDATE SET km=excluded.km",
            (o.year, o.plate.strip(), o.km))
        conn.commit()
    return get_odometer(year=o.year, plate=o.plate)


# --------------------------------------------------------------- export

def _book_rows(from_ts, to_ts, plate):
    lo, hi = ts_range(from_ts, to_ts)
    psql, pargs = _plate_sql(plate)
    with closing(db.connect()) as conn:
        return conn.execute(
            f"SELECT * FROM trips WHERE start_ts BETWEEN ? AND ? AND excluded=0{psql} "
            f"ORDER BY start_ts", (lo, hi, *pargs)).fetchall()


@router.get("/export.xlsx")
def export_spz(from_ts: int | None = Query(None), to_ts: int | None = Query(None),
               plate: str | None = Query(None)):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows = _book_rows(from_ts, to_ts, plate)

    wb = Workbook()
    ws = wb.active
    ws.title = "Kniha jízd"
    headers = ["SPZ", "Datum", "Odjezd", "Příjezd", "Odkud", "Kam",
               "Účel jízdy", "Km", "Řidič", "Soukromá"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        s, e = fmt_dt(r["start_ts"]), fmt_dt(r["end_ts"])
        ws.append([r["plate"], s.date(), s.strftime("%H:%M"), e.strftime("%H:%M"),
                   r["origin"], r["destination"], r["purpose"], r["km"],
                   r["driver"], "ano" if r["private"] else "ne"])
    for col, w in zip("ABCDEFGHIJ", [12, 11, 8, 8, 26, 26, 22, 7, 16, 9], strict=True):
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        row[0].number_format = "DD.MM.YYYY"
    ws.freeze_panes = "A2"

    total = ws.max_row + 1
    ws.cell(row=total, column=7, value="Celkem").font = Font(bold=True)
    ws.cell(row=total, column=8,
            value=round(sum(r["km"] for r in rows), 1)).font = Font(bold=True)
    return xlsx_response(wb, "kniha-jizd-spz.xlsx")


_FONT_DIRS = ("/usr/share/fonts/truetype/dejavu", "/usr/share/fonts/dejavu")


def _pdf_fonts() -> tuple[str, str]:
    """Zaregistruje TTF s českou diakritikou; fallback na Helveticu."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    for d in _FONT_DIRS:
        reg = os.path.join(d, "DejaVuSans.ttf")
        bold = os.path.join(d, "DejaVuSans-Bold.ttf")
        if os.path.exists(reg) and os.path.exists(bold):
            pdfmetrics.registerFont(TTFont("Deja", reg))
            pdfmetrics.registerFont(TTFont("DejaB", bold))
            return "Deja", "DejaB"
    return "Helvetica", "Helvetica-Bold"


@router.get("/export.pdf")
def export_pdf(from_ts: int | None = Query(None), to_ts: int | None = Query(None),
               plate: str | None = Query(None), driver: str = Query("")):
    """Kniha jízd jako PDF – pro tisk a předání účetní."""
    import io

    from fastapi.responses import Response as FResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    rows = _book_rows(from_ts, to_ts, plate)
    font, font_b = _pdf_fonts()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm,
                            title="Kniha jízd")
    h1 = ParagraphStyle("h1", fontName=font_b, fontSize=15, spaceAfter=2)
    meta = ParagraphStyle("meta", fontName=font, fontSize=9.5, textColor=colors.grey)

    period = ""
    if rows:
        period = (f"{fmt_dt(rows[0]['start_ts']):%d.%m.%Y} – "
                  f"{fmt_dt(rows[-1]['start_ts']):%d.%m.%Y}")
    plates = sorted({(r["plate"] or "").strip() for r in rows} - {""})
    drivers = sorted({(r["driver"] or "").strip() for r in rows} - {""})
    story = [
        Paragraph("Kniha jízd", h1),
        Paragraph(" · ".join(filter(None, [
            f"Vozidlo: {', '.join(plates) or '–'}",
            f"Řidič: {driver or ', '.join(drivers) or '–'}",
            f"Období: {period or '–'}"])), meta),
        Spacer(0, 5 * mm),
    ]

    header = ["Datum", "Odjezd", "Příjezd", "Odkud", "Kam", "Účel jízdy",
              "Km", "SPZ", "Soukr."]
    data = [header]
    style = [
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("FONTNAME", (0, 0), (-1, 0), font_b),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (6, 0), (6, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, 0), 0.75, colors.black),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.Color(0.96, 0.96, 0.95)]),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]
    month = None
    month_km = 0.0
    total_km = 0.0

    def close_month():
        nonlocal month_km
        if month is not None:
            data.append(["", "", "", "", "", f"Součet {month}",
                         f"{month_km:g}", "", ""])
            r = len(data) - 1
            style.append(("FONTNAME", (0, r), (-1, r), font_b))
            style.append(("LINEABOVE", (0, r), (-1, r), 0.4, colors.grey))
            month_km = 0.0

    for r in rows:
        s, e = fmt_dt(r["start_ts"]), fmt_dt(r["end_ts"])
        m = s.strftime("%m/%Y")
        if month is not None and m != month:
            close_month()
        month = m
        data.append([s.strftime("%d.%m.%Y"), s.strftime("%H:%M"), e.strftime("%H:%M"),
                     r["origin"] or "", r["destination"] or "", r["purpose"] or "",
                     f"{r['km']:g}", r["plate"] or "", "ano" if r["private"] else ""])
        month_km += r["km"]
        total_km += r["km"]
    close_month()
    data.append(["", "", "", "", "", "CELKEM", f"{round(total_km, 1):g}", "", ""])
    style.append(("FONTNAME", (0, len(data) - 1), (-1, len(data) - 1), font_b))
    style.append(("LINEABOVE", (0, len(data) - 1), (-1, len(data) - 1), 0.9, colors.black))

    table = Table(data, repeatRows=1, colWidths=[
        22 * mm, 15 * mm, 15 * mm, 52 * mm, 52 * mm, 52 * mm, 14 * mm, 24 * mm, 14 * mm])
    table.setStyle(TableStyle(style))
    story.append(table)
    doc.build(story)
    return FResponse(buf.getvalue(), media_type="application/pdf",
                     headers={"Content-Disposition": 'attachment; filename="kniha-jizd.pdf"'})
